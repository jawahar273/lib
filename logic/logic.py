'''
Created on 03-Jun-2016

@author: jon
'''

from openpyxl import Workbook
from openpyxl import load_workbook
from os import popen
from PyQt4.QtGui import QMessageBox, QMainWindow, QTableWidgetItem, QInputDialog, QWidget, QRadioButton
from PyQt4 import QtCore

from view.mainwindow import *

try:
    _fromUtf8 = QtCore.QString.fromUtf8
    QString = QtCore.QString
except AttributeError:
    def _fromUtf8(s):
        return s



class ThreadLib(QtCore.QThread):
    def __init__(self,f):
        QtCore.QThread.__init__(self)
        self.fun = f

    def run(self):
       self.fun()

    def __del__(self):
    	self.wait()

class PopWidget(QWidget):
    def __init__(self):
        QWidget.__init__(self)
        self.text, self.ok = QInputDialog.getText(self,
                                                  "Name of the sheet",
                                                  "Enter the correct name of current sheet to access",
                                                  )

    def text_r(self):
        if self.ok:
            return self.text


#################################################################################
#
# maseter unique cell must be set carfully
# :variable:::maseter_head
#
#################################################################################
class Main(QMainWindow):
    '''
    classdocs
    '''

    def __init__(self, parent=None):
        '''
        Constructor
        '''
        QMainWindow.__init__(self, parent)
        self.load_wb = None
        self.load_wb2 = None
        self.select = None
        self.select2 = None
        self.master_set = list()
        self.slave_set = list()
        self.len_row = 0
        self.len_row2 = 0
        self.len_col = 0
        self.len_col2 = 0
        self.row_1 = 2
        self.row_data = None
        self.len_col = None
        self.len_row_str = " "
        self.len_row2_str = " "
        self.len_col_str = " "
        self.master_header = "AccessNo"
        self.len_col2_str = " "
        self.ok_2 = "Success"
        self.ok_1 = "Success"

        self.j_loop = 1
        self.col_i = 1
        # self.Range = None
        self.ui_window = Ui_MainWindow()
        self.ui_window.setupUi(self)


        self.wb = Workbook()
        self.wb_sheet1_1 = self.wb.active
        #)

        #self.gridLayout_5.addWidget(self.radioButton_2
        # self.select2['A1:A'+self.len_row2]

        # row_count_slave = 1
        # row_count_master = 1

        # self.search_in_master()


    def load_file_1(self, a):
        """
        this file is used to load file Master excel file
        :param a: file name
        :return: cuts the file and take the value in Master excel file
        """

        # a = "Master.xlsx"
        self.load_wb = load_workbook(a)

        if(self.master_set):
        	self.master_set.remove()

        try:
            self.__u = self.load_wb.get_sheet_names()[0]

            self.select = self.load_wb[self.__u]

        except KeyError:
            try:
                self.select = self.load_wb["Sheet"]
            except KeyError:

                self.select = self.load_wb[_fromUtf8(PopWidget().text_r())]
            else:
                self.ok_1 = "fails"

        self.len_row = (len(self.select.rows))
        self.len_row_str = str(self.len_row)
        # start
        self.len_col = len(self.select.columns)
        self.len_col_str = str(self.len_col)
        ####print(("No of columns in Master"+self.len_col)
        # end
        self.l = (self.select['A1:A' + self.len_row_str])
        count = 1
        for j in self.l:
            self.master_set.append(self.select['A' + str(count)].value)
            ####print((self.select['A'+str(count)].value)
            count += 1

    def load_file_2(self, a):
        """
        this file is used to load file slave excel file
        :param a: file name
        :return: cuts the file and take the value in slave excel file
        """

        # a = "Slave1.xlsx"
        self.load_wb2 = load_workbook(a)

        if(self.slave_set):
        	self.slave_set.remove()

        try:
            self.select2 = self.load_wb2["Sheet1"]
        except KeyError:
            try:
                self.select2 = self.load_wb2["Sheet"]
            except KeyError:

                self.select2 = self.load_wb2[_fromUtf8(PopWidget().text_r())]
            else:
                self.ok_2 = "fails"

        self.len_row2 = len(self.select2.rows)
        self.len_row2_str = str(self.len_row2)
        count = 1
        for slave in self.select2['A1:A' + self.len_row2_str]:
            self.slave_set.append(self.select2['A' + str(count)].value)
            ####print((">>>>slave",(self.select2['A'+str(count)].value))
            count += 1
            # self.store_set = list(0)

    def ok_text_1(self):
        return self.ok_1

    def ok_text_2(self):
        return self.ok_2

    def save_file(self, a):
        """
        This function is used to save the file in .xlsx format
        :param a: file name
        :return: save the excel file with the given name
        """

        try:

            #    self.store.remove(None)
            #  self.store.remove(self.master_header)
            ##print((self.store)
            self.slave_set.remove(None)
            self.slave_set.remove(self.master_header)
        except ValueError:
            pass
        try:
        	popen("del"+a)
        except FileNotFoundError:
        	pass

        """
        row_list = []                  # start
        for i in range(1,self.len_col):
           row_data = self.row_values(i)
           row_list.append(row_data)

                                       # end
        """

        self.master_set = list(self.master_set)

        # print((self.master_set)
        try:

            self.master_set.remove(None)
            self.master_set.remove(self.master_header)
        except ValueError as v:
            print("V-->\n", v)

        self.len_row = len(self.master_set)
        # print(("master set", self.len_row, "\n row len of file1:", self.len_row, '\n', "====" * 10)

        try:
            self.master_set.sort()

        except TypeError:
            # print(self.master_set,"slave sort:",self.slave_set)
            pass

            # print(("master set", "--" * 20, len(self.master_set), "\n", "--" * 20, "\n len of slave", len(self.slave_set))

            # <-----------------------dont delect this



        self.col_i = 0

        # print(self.select.rows, self.select.columns)
        try:
          self.ui_window.tableWidget.setRowCount(len(self.select.rows))
          self.ui_window.tableWidget.setColumnCount(len(self.select.columns))
        except AttributeError:
            pass
        print((type(self.len_row),self.len_row, "==",len(self.select.rows)))

        ThreadLib(self.same_check).start()

        print(self.row_1)
        print("cure",self.col_i)
        try:
            self.wb.save(a)

            # print(()
        except PermissionError:
            # #print(("close the report file please..")
            '''
            from os import system
            system("taskkill /IM excel.exe")
            '''
            QMessageBox.information(None, "version", _fromUtf8("closing result  files\n writeing is not allowed on opened file then open the file..."))#
            self.wb.save(a)


    def missing_check(self):
      try:

            for i in self.select["A1:I1"]:
                """
            --- For inserting header on result excel from master excel
             """
                coutn = 1

                for j in i:
                    self.wb_sheet1_1.cell(row=1, column=coutn, value=j.value)
                    coutn += 1
      except TypeError as t:
            print("T-----------------\n", t)
      self.col_i =0
      g = 0
      enable = False

      while(g < self.len_row):
        try:

                a1 = self.slave_set[self.col_i]
        except IndexError:
                print("slave Index error ====toS: ", toS, "self.col_i:", self.col_i)


        if(enable == True):
               g = g - 1
        b1 = self.master_set[g]
  
        
        g+=1
        toS= str(g)
        if (a1 != b1):
            print("slave:", a1, "!=", b1, ":master")
            
            enable = False
            if(a1 < b1):
                  print("\t \t \t \t \t \t \t  addition values in slave:", a1)
                  self.col_i += 1
                  enable = True
            else:
            	self.compare_to_check(toS)
        else:
                self.col_i += 1
                enable = False
       


    def same_check(self):
        try:

            for i in self.select["A1:I1"]:
                """
            --- For inserting header on result excel from master excel
             """
                coutn = 1

                for j in i:
                    self.wb_sheet1_1.cell(row=1, column=coutn, value=j.value)
                    coutn += 1
        except TypeError as t:
            print("T-----------------\n", t)
        enable = False

        g = 0
        self.col_i =0
        while(g < self.len_row):
            """"
            :: self.col_i == counting slave loop
            """

            try:

                a1 = self.slave_set[self.col_i]
            except IndexError:
                print("Index error ====G: ", g, "self.col_i:", self.col_i)
                break

            if(enable == True):
               g = g - 1
            b1 = (self.master_set[g])
            g+=1
            toS= str(g)
            if(self.len_row == self.col_i):
                break
            if (a1 == b1):
              print(   "master:", b1, "==", a1,":slave" )
              self.compare_to_check(toS)
              print("index for master file:",g,"index for slave file",self.col_i,"(toS):", toS)
              self.col_i += 1
            else:
                print(" \t \t \t \t \t \t \t slave:", a1, "!=", b1, ":master")
                enable = False

                if(a1 < b1):
                  print("\t \t \t \t \t \t \t  addition values in slave:", a1)
                  self.col_i += 1
                  enable = True

    def compare_to_check(self, toS):
        """

        :type self: object
        :type toS: int(value) to string
        """
        to = self.select["A" + toS + ":I" + toS]
        # print(("row count in master:",toS)
        # print((to ,"\n rotation:",self.col_i)

        self.ui_window.tableWidget.setColumnCount(self.len_col)
        self.ui_window.tableWidget.setRowCount(self.row_1)
        for r_o in to:
            col_1 = 1
            #print(("row---->",self.row_1,"\ncol----------------------------------------------------")
            for c_o in r_o:
                try:

                      
                      self.ui_window.tableWidget.setItem(self.row_1-1, col_1-1, QTableWidgetItem(str(c_o.value)))
                except (TypeError):
                    #print("...........")
                    self.ui_window.tableWidget.setItem(self.row_1, col_1-1, QTableWidgetItem(QtCore.QDate.toString(c_o.value)))
                  
                self.wb_sheet1_1.cell(row=self.row_1, column=col_1, value=c_o.value)
                ##print((c_o.value)
                col_1 += 1
            ###print(()
            self.row_1 += 1

